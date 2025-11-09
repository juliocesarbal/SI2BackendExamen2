"""
Voice Assistant Helper Functions
Generates reports in multiple formats (PDF, Excel, HTML)
"""

import json
import base64
from io import BytesIO
from datetime import datetime, timedelta, date
from django.db.models import F
from decouple import config
import os


def transcribe_audio_with_gemini(audio_base64: str, mime_type: str):
    """Transcribe audio using Gemini AI"""
    try:
        from google import generativeai as genai

        api_key = config('GEMINI_API_KEY', default='')
        if not api_key:
            raise ValueError('GEMINI_API_KEY no está configurada en las variables de entorno')

        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.0-flash-exp')

        prompt = """Transcribe el siguiente audio en español. El audio contiene un comando para generar reportes en un sistema de farmacia.
Devuelve SOLO el texto transcrito, sin explicaciones adicionales."""

        result = model.generate_content([
            {
                'inline_data': {
                    'data': audio_base64,
                    'mime_type': mime_type
                }
            },
            prompt
        ])

        transcript = result.text.strip()
        print(f'[Gemini] Audio transcrito: {transcript}')
        return transcript

    except Exception as e:
        print(f'Error transcribing audio: {e}')
        raise


def parse_command_with_gemini(command: str):
    """Parse voice command using Gemini AI"""
    try:
        system_prompt = f"""Eres un asistente inteligente para el panel de administración de una farmacia.
Tu tarea es analizar comandos de voz (que pueden tener errores de transcripción) y extraer la información necesaria para generar reportes.

IMPORTANTE: Debes ser muy tolerante a errores de transcripción y usar contexto para corregir automáticamente.

Los tipos de reportes disponibles son:
1. ALERTAS - Reportes de inventario bajo stock
2. BITACORA - Registros de seguridad del sistema
3. CLIENTES - Información de clientes registrados
4. FACTURAS - Reportes de ventas y facturas

Los formatos disponibles son: PDF, EXCEL, HTML
- Si el usuario no especifica formato, usa PDF por defecto

FILTROS DE FECHA:
- Si el usuario NO especifica fechas o intervalo: deja "fechaInicio" y "fechaFin" como null (se exportarán TODOS los registros)
- Si el usuario menciona "último mes" o "mes pasado": usa el valor "último mes" para fechaInicio
- Si el usuario menciona "última semana" o "semana pasada": usa el valor "última semana" para fechaInicio

Responde SOLO con un objeto JSON con la siguiente estructura:
{{
  "action": "generar_reporte",
  "reportType": "ALERTAS|BITACORA|CLIENTES|FACTURAS",
  "format": "PDF|EXCEL|HTML",
  "filters": {{
    "fechaInicio": null,
    "fechaFin": null,
    "tipo": null
  }}
}}

Comando del usuario: "{command}"
"""

        from google import generativeai as genai

        api_key = config('GEMINI_API_KEY', default='')
        if not api_key:
            raise ValueError('GEMINI_API_KEY no está configurada en las variables de entorno')

        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.0-flash-exp')

        result = model.generate_content(system_prompt)
        text = result.text

        # Extract JSON from response
        start = text.find('{')
        end = text.rfind('}') + 1
        if start == -1 or end == 0:
            raise ValueError('No se pudo parsear el comando')

        json_text = text[start:end]
        return json.loads(json_text)
    except Exception as e:
        print(f'Error parsing command: {e}')
        raise


def process_date_filters(filters):
    """Process date filters from voice command"""
    processed = filters.copy()

    # If no dates specified, leave as null to get ALL records
    if not filters.get('fechaInicio') and not filters.get('fechaFin'):
        processed['fechaInicio'] = None
        processed['fechaFin'] = None
        return processed

    # Handle "último mes"
    if filters.get('fechaInicio') == 'último mes':
        today = datetime.now().date()
        last_month = date(today.year, today.month - 1, 1) if today.month > 1 else date(today.year - 1, 12, 1)
        processed['fechaInicio'] = last_month.isoformat()
        processed['fechaFin'] = today.isoformat()

    # Handle "última semana"
    if filters.get('fechaInicio') == 'última semana':
        today = datetime.now().date()
        last_week = today - timedelta(days=7)
        processed['fechaInicio'] = last_week.isoformat()
        processed['fechaFin'] = today.isoformat()

    return processed


def generate_pdf(title, data, columns, timestamp):
    """Generate PDF report using reportlab"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#6366f1'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        elements.append(Paragraph(title.replace('_', ' '), title_style))

        # Date
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            spaceAfter=20,
            alignment=TA_CENTER
        )
        elements.append(Paragraph(f'Fecha: {datetime.now().strftime("%d/%m/%Y")}', date_style))
        elements.append(Spacer(1, 0.2*inch))

        # Table
        table_data = [[col['header'] for col in columns]]
        for row in data:
            table_data.append([str(row.get(col['key'], '')) for col in columns])

        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))

        elements.append(table)
        doc.build(elements)

        pdf_data = buffer.getvalue()
        buffer.close()

        return {
            'fileData': base64.b64encode(pdf_data).decode('utf-8'),
            'fileName': f'{title}_{timestamp}.pdf',
            'mimeType': 'application/pdf'
        }
    except Exception as e:
        print(f'Error generating PDF: {e}')
        raise


def generate_excel(title, data, columns, timestamp):
    """Generate Excel report using openpyxl"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name limit

        # Header row
        header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col['header'])
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_idx, row in enumerate(data, 2):
            for col_idx, col in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col['key'], ''))

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        buffer = BytesIO()
        wb.save(buffer)
        excel_data = buffer.getvalue()
        buffer.close()

        return {
            'fileData': base64.b64encode(excel_data).decode('utf-8'),
            'fileName': f'{title}_{timestamp}.xlsx',
            'mimeType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
    except Exception as e:
        print(f'Error generating Excel: {e}')
        raise


def generate_html(title, data, columns, timestamp):
    """Generate HTML report"""
    html = f"""
<!DOCTYPE html>
<html>
<head>
  <meta charset="UTF-8">
  <title>{title.replace('_', ' ')}</title>
  <style>
    body {{
      font-family: Arial, sans-serif;
      margin: 20px;
      background-color: #f5f5f5;
    }}
    .container {{
      max-width: 1200px;
      margin: 0 auto;
      background-color: white;
      padding: 30px;
      border-radius: 8px;
      box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }}
    h1 {{
      color: #6366f1;
      margin-bottom: 10px;
    }}
    .date {{
      color: #666;
      margin-bottom: 20px;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      margin-top: 20px;
    }}
    th {{
      background-color: #6366f1;
      color: white;
      padding: 12px;
      text-align: left;
      font-weight: 600;
    }}
    td {{
      padding: 10px;
      border-bottom: 1px solid #e5e7eb;
    }}
    tr:hover {{
      background-color: #f9fafb;
    }}
    .footer {{
      margin-top: 30px;
      padding-top: 20px;
      border-top: 1px solid #e5e7eb;
      color: #666;
      font-size: 14px;
      text-align: center;
    }}
  </style>
</head>
<body>
  <div class="container">
    <h1>{title.replace('_', ' ')}</h1>
    <div class="date">Fecha: {datetime.now().strftime('%d/%m/%Y')}</div>
    <table>
      <thead>
        <tr>
          {''.join([f'<th>{col["header"]}</th>' for col in columns])}
        </tr>
      </thead>
      <tbody>
        {''.join([
            '<tr>' +
            ''.join([f'<td>{row.get(col["key"], "")}</td>' for col in columns]) +
            '</tr>'
            for row in data
        ])}
      </tbody>
    </table>
    <div class="footer">
      <p>Total de registros: {len(data)}</p>
      <p>Generado el {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</p>
    </div>
  </div>
</body>
</html>
"""

    return {
        'fileData': base64.b64encode(html.encode('utf-8')).decode('utf-8'),
        'fileName': f'{title}_{timestamp}.html',
        'mimeType': 'text/html'
    }


def generate_file(title, format_type, data, columns):
    """Generate file in specified format"""
    timestamp = datetime.now().strftime('%Y-%m-%d')

    if format_type == 'PDF':
        return generate_pdf(title, data, columns, timestamp)
    elif format_type == 'EXCEL':
        return generate_excel(title, data, columns, timestamp)
    elif format_type == 'HTML':
        return generate_html(title, data, columns, timestamp)
    else:
        raise ValueError(f'Formato no soportado: {format_type}')
